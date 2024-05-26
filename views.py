import os
import openpyxl as xl
import mysql.connector as a
import easyocr
import pytesseract,re
import datetime ,os
from django.http import HttpResponse
from django.shortcuts import render,redirect
from django.contrib.auth import login  , authenticate 
from django.contrib.auth.forms import AuthenticationForm
from forms import SignUpForm, LoginForm
def signup_view(request):
    if request.method == 'POST':
        form = SignUpForm(request.POST)
        if form.is_valid():
            form.save()
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password1')
            user = authenticate(username=username, password=password)
            login(request, user)
            return render(request,'Attendancemainpage.html')
    else:
        form = SignUpForm()
    return render(request, 'signup.html', {'form': form})
def login_view(request):
    if request.method == 'POST':
        form = LoginForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                params = {'signup':request.POST.get("username")}
                return render(request,'Attendancemainpage.html' , params)
    else:
        form = LoginForm()
    return render(request, 'login.html', {'form': form})
# def loginPage(request):
#     return render(request , 'index.html')
def Contactus(request):
    return render(request , 'Contact.html')
def index(request):
    params = {'signup':request.GET.get()}
    return render(request ,'Attendancemainpage.html',params)
def submit_form(request):
    params = {'signup':request.GET.get("email")}
    return render(request , 'SubmissionForm.html' , params)

#def your_view_function(file_name):
    #module_dir = os.path.dirname(__file__)   #get current directory
    #file_path = os.path.join(module_dir, file_name)   #full path to text.
    
def convert_string_to_int(string_number):
    # Remove any spaces from the string
    cleaned_string = string_number.replace(" ", "")
    cleaned_string = cleaned_string.replace("_" ,"")
    
    # Ensure the cleaned string only contains digits before conversion
    if cleaned_string.isdigit():
        integer_number = int(cleaned_string)
        return integer_number
    else:
        raise ValueError("The input string contains non-digit characters")
def Excelwork(branch , subject, xlfile , date , time, request):
    wb = xl.load_workbook(xlfile)
    ws = wb['Attendance_sheet']
    ws.cell(row = 1 , column=2).value = 'SName'
    ws.cell(row = 1 , column=3).value = 'CollegeID'
            # matches = re.findall(r'[0-9][0-9][0-9]',text)
    IMAGE_PATH = request.GET.get("filechoosing","Default")
    reader = easyocr.Reader(['en'],gpu=True)
    results = reader.readtext(IMAGE_PATH)
    text = ''
    matches = []
    for result in results:
                text =result[1]
                matches.append(text)
    count = 2
    while(ws.cell(row=count,column=3).value):
            count=count+1
    x = 4
    present = 0 
    absent = 0 
    while(x < 100):
                if(ws.cell(row = 1 ,column=x).value == None ):
                    ws.cell(row= 1 , column=x).value =f'{date}'
                    for i in range(2,count):
                        for id in matches:
                            idin = ws.cell(row=i,column=3).value
                            # newid =int(id)
                            newid =  convert_string_to_int(id)
                            if(newid == idin):
                                ws.cell(row = i , column=x).value="Present"
                                present=present+1
                    i=count-1
                    while(i!=1):
                        t = ws.cell(row=i ,column=x).value
                        if(t == None):
                            ws.cell(row= i ,column=x).value = "Absent"
                            absent =absent + 1
                        i=i-1
                    ws.cell(row = count , column = x).value = time
                    ws.cell(row = count+2 , column =x ).value = present 
                    ws.cell(row = count+3 , column =x).value = absent 
                    attendance_per = (present/(absent+present))*100
                    ws.cell(row=count+4 , column = x ).value = attendance_per
                    ws.cell(row = count+1, column= x).value =subject
                    break
                else:
                    x=x+1  
    wb.save(xlfile)   
def takeinputs(request):
    date = datetime.date.today()
    time = request.GET.get("time","Default")
    section = request.GET.get("section","Default")
    subject = request.GET.get("subject","Default")
    Branch = request.GET.get("branchName","Default")
    year = request.GET.get("year","Default")
    date = request.GET.get("date","Default")
    hostname = request.GET.get("")
    if(Branch == "CS"):
        if(section == 'B'):
            Excelwork(Branch,subject,"B.xlsx",date ,time,request)
        elif(section == 'C'):
            Excelwork(Branch,subject,"Cyber.xlsx",date,time,request)
        elif(section == 'A'):
            Excelwork(Branch,subject ,"A.xlsx",date,time,request)
    elif(Branch == "EC"):
        Excelwork(Branch,subject, "E.xlsx",date ,time , request)
    elif(Branch == "ME"):
        Excelwork(Branch , subject ,"Me.xlsx" , date, time, request)
    elif(Branch == "IT"):
        Excelwork(Branch ,subject ,"I.xlsx", date , time, request )
    elif(Branch == "CIVIL"):
        Excelwork(Branch , subject , "C.xlsx" , date ,time, request)
    # loading an excel file 
           
    # con =a.connect(host="127.0.0.1",user="root",passwd="supersaiyan1000")
    # c = con.cursor()
    # c.execute("show databases")
    # dl = c.fetchall()
    # f= False
    # for i in dl: 
    #     if 'arbitary'==i[0]:
    #         sql = "use Arbitary"
    #         c.execute(sql)
    #         f=True
    # tablecreated = False
    # if f==False:
    #     sql1="create database Arbitary"
    #     c.execute(sql1)
    # c.execute("show tables")
    # dl = c.fetchall()
    # for i in dl:
    #     if('attendance_tablesheet' == i[0]):
    #         tablecreated = True
    # if(tablecreated == False):
    #         sql3 = "create table Attendance_tablesheet(Name char(15),Rollno varchar(10) , Attendance char(10) , dateofclass varchar(20))"
    #         c.execute(sql3)
    #         tablecreated = True
    # dateofclass = date
    # for i in range(2,11):
    #         idinName = ws.cell(row= i , column=2).value
    #         idinRoll = ws.cell(row= i , column=3).value
    #         idinatt = ws.cell(row = i , column=4).value 
    #         sql4 = f"insert into Attendance_tablesheet values('{idinName}',{idinRoll},'{idinatt}','{dateofclass}')"
    #         c.execute(sql4)
    # sql = "insert into attendance_tablesheet values('-----','-----','------','------')"
    # c.execute(sql)
    # con.commit()
    # con.close
    
def thankyou(request):
    params = {'section' : request.GET.get("section"),
              'year' : request.GET.get("year"),
              'branchName' : request.GET.get("branchName"),
              'totalp' : request.GET.get("totalp"),
              'date' : request.GET.get("date"),
              'time ':request.GET.get("time"),
              'subject ': request.GET.get("subject")} 
    takeinputs(request)
    return render(request , 'thankyoupage.html',params )

